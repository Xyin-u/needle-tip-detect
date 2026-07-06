import json
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
import torch
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from torchvision.ops import nms
from tqdm import tqdm

from ultralytics import YOLO


# -------------------------------------------------------------------
# 1. 滑动窗口检测函数（保持不变）
# -------------------------------------------------------------------
def sliding_window_pose(
    image_paths,
    model,
    window_size=(480, 480),
    step_x=80,
    step_y=80,
    batch_size=16,
    conf_thresh=0.25,
    iou_thresh=0.5,
):
    win_w, win_h = window_size
    model_input_size = model.overrides.get("imgsz", 480)
    num_kpts = model.model.model[-1].kpt_shape[0] if hasattr(model.model.model[-1], "kpt_shape") else 1

    all_results = []
    for img_path in tqdm(image_paths, desc="Processing images"):
        img = cv2.imread(str(img_path))
        if img is None:
            continue
        h, w = img.shape[:2]

        windows = []
        for y in range(0, h, step_y):
            for x in range(0, w, step_x):
                x_end = min(x + win_w, w)
                y_end = min(y + win_h, h)
                windows.append((x, y, x_end, y_end))

        all_dets = []

        for batch_start in range(0, len(windows), batch_size):
            batch_wins = windows[batch_start : batch_start + batch_size]
            batch_crops = []
            batch_offsets = []
            for x1, y1, x2, y2 in batch_wins:
                crop = img[y1:y2, x1:x2]
                batch_crops.append(crop)
                batch_offsets.append((x1, y1))

            results = model(batch_crops, imgsz=model_input_size, conf=conf_thresh, verbose=False)

            for res, (x_off, y_off) in zip(results, batch_offsets):
                if res.boxes is None:
                    continue
                boxes = res.boxes.xyxy.cpu().numpy()
                confs = res.boxes.conf.cpu().numpy()
                clss = res.boxes.cls.cpu().numpy().astype(int)
                kpts_xy = res.keypoints.xy.cpu().numpy() if res.keypoints is not None else None
                kpts_conf = res.keypoints.conf.cpu().numpy() if res.keypoints is not None else None

                for i in range(len(boxes)):
                    x1_abs = max(0, min(w, x_off + boxes[i][0]))
                    y1_abs = max(0, min(h, y_off + boxes[i][1]))
                    x2_abs = max(0, min(w, x_off + boxes[i][2]))
                    y2_abs = max(0, min(h, y_off + boxes[i][3]))

                    if kpts_xy is None or i >= len(kpts_xy):
                        continue
                    if kpts_conf is None:
                        continue

                    kpts_abs = []
                    for j in range(num_kpts):
                        kx = kpts_xy[i][j][0]
                        ky = kpts_xy[i][j][1]
                        kc = kpts_conf[i][j]
                        if kx == 0 and ky == 0 and kc == 0:
                            kpts_abs.append((0.0, 0.0, 0.0))
                        else:
                            kx_abs = x_off + kx
                            ky_abs = y_off + ky
                            kx_abs = max(0, min(w, kx_abs))
                            ky_abs = max(0, min(h, ky_abs))
                            kpts_abs.append((kx_abs, ky_abs, kc))

                    if not kpts_abs or len(kpts_abs) == 0:
                        continue
                    kpt_conf = kpts_abs[0][2]
                    if kpt_conf < 0.3:
                        continue

                    all_dets.append(
                        {
                            "bbox": [x1_abs, y1_abs, x2_abs, y2_abs],
                            "bbox_conf": float(confs[i]),
                            "class_id": int(clss[i]),
                            "keypoints": kpts_abs,
                        }
                    )

        if all_dets:
            boxes_tensor = torch.tensor([d["bbox"] for d in all_dets])
            scores_tensor = torch.tensor([d["bbox_conf"] for d in all_dets])
            keep = nms(boxes_tensor, scores_tensor, iou_thresh)
            nms_dets = [all_dets[i] for i in keep.tolist()]

            if nms_dets:
                best_det = max(nms_dets, key=lambda d: d["bbox_conf"])
                final_dets = [best_det]
            else:
                final_dets = []
        else:
            final_dets = []

        all_results.append({"img_path": str(img_path), "detections": final_dets, "num_keypoints": num_kpts})
    return all_results


# -------------------------------------------------------------------
# 2. 加载真实标签（保持不变）
# -------------------------------------------------------------------
def load_groundtruth(img_path, gt_dir):
    img_stem = Path(img_path).stem
    json_path = Path(gt_dir) / f"{img_stem}.json"
    if not json_path.exists():
        return []

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    needle_box = None
    tip_point = None
    for shape in data.get("shapes", []):
        label = shape.get("label", "")
        if label == "needle" and shape.get("shape_type") == "rectangle":
            points = shape["points"]
            xs = [p[0] for p in points]
            ys = [p[1] for p in points]
            x1, x2 = min(xs), max(xs)
            y1, y2 = min(ys), max(ys)
            needle_box = [x1, y1, x2, y2]
        elif label == "tip" and shape.get("shape_type") == "point":
            tip_point = shape["points"][0]

    if needle_box is None or tip_point is None:
        return []

    keypoints = [(tip_point[0], tip_point[1])]
    return [{"bbox": needle_box, "keypoints": keypoints}]


# -------------------------------------------------------------------
# 3. 绘制结果图（保持不变）
# -------------------------------------------------------------------
def draw_results(img, dets, gts, output_path, show_coords=True, show_conf=True):
    img_copy = img.copy()
    for gt in gts:
        for x, y in gt["keypoints"]:
            if x != 0 or y != 0:
                cv2.circle(img_copy, (int(x), int(y)), 1, (255, 0, 0), -1)
                if show_coords:
                    text = f"({int(x)},{int(y)})"
                    cv2.putText(img_copy, text, (int(x) + 5, int(y) - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 0, 0), 1)
    for det in dets:
        x1, y1, x2, y2 = map(int, det["bbox"])
        conf = det["bbox_conf"]
        cv2.rectangle(img_copy, (x1, y1), (x2, y2), (0, 255, 0), 1)
        if show_conf:
            conf_text = f"{conf:.2f}"
            text_x, text_y = x1, y1 - 5
            if text_y < 5:
                text_y = y1 + 15
            cv2.putText(img_copy, conf_text, (text_x, text_y), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)
        for x, y, conf_kpt in det["keypoints"]:
            if conf_kpt > 0.3 and (x != 0 or y != 0):
                cv2.circle(img_copy, (int(x), int(y)), 1, (0, 0, 255), -1)
                if show_coords:
                    text = f"({int(x)},{int(y)})"
                    cv2.putText(
                        img_copy, text, (int(x) + 5, int(y) + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 255), 1
                    )
    cv2.imwrite(output_path, img_copy)


# -------------------------------------------------------------------
# 4. 主函数（修改：删除误检/漏检列，保留末尾统计）
# -------------------------------------------------------------------
def main():
    # ========== 用户配置区域 ==========
    img_folder = r"C:\Users\tiger\Desktop\tip_dataset\testdata\test0608"
    gt_folder = r"C:\Users\tiger\Desktop\tip_dataset\testdata\test0608"
    output_folder = r"D:\xinyu\ultralytics-main\runs\pose\results26\0608_240_01"
    model_path = r"D:\xinyu\ultralytics-main\runs\pose\tip_pose扩增_26pose\100e_0609_01\weights\best.pt"
    window_size = (480, 480)
    step_x = 240
    step_y = 240
    batch_size = 16
    conf_thresh = 0.5
    iou_nms = 0.2
    # ==================================

    out_img_dir = Path(output_folder) / "visualizations"
    out_img_dir.mkdir(parents=True, exist_ok=True)
    out_excel = Path(output_folder) / "evaluation_results.xlsx"

    model = YOLO(model_path)

    img_paths = list(Path(img_folder).glob("*.jpg")) + list(Path(img_folder).glob("*.png"))
    img_paths = [str(p) for p in img_paths]
    print(f"找到 {len(img_paths)} 张图片")

    det_results = sliding_window_pose(
        image_paths=img_paths,
        model=model,
        window_size=window_size,
        step_x=step_x,
        step_y=step_y,
        batch_size=batch_size,
        conf_thresh=conf_thresh,
        iou_thresh=iou_nms,
    )

    per_image_records = []
    all_distances = []
    false_positive_count = 0  # 统计误检个数
    miss_count = 0  # 统计漏检个数

    for res in tqdm(det_results, desc="Evaluating"):
        img_path = res["img_path"]
        img_name = Path(img_path).stem
        img = cv2.imread(img_path)
        if img is None:
            continue

        gts = load_groundtruth(img_path, gt_folder)
        dets = res["detections"]

        gt_x, gt_y = None, None
        if gts and len(gts) > 0 and len(gts[0]["keypoints"]) > 0:
            gt_x, gt_y = gts[0]["keypoints"][0]

        det_x, det_y = None, None
        has_tip = False
        if dets and len(dets) > 0:
            kpts = dets[0]["keypoints"]
            if kpts and len(kpts) > 0:
                dx, dy, conf = kpts[0]
                if conf > 0.3 and (dx != 0 or dy != 0):
                    det_x, det_y = dx, dy
                    has_tip = True

        distance = None
        if has_tip and gt_x is not None and gt_y is not None:
            distance = np.hypot(det_x - gt_x, det_y - gt_y)
            all_distances.append(distance)

        # 判断误检和漏检（仅用于计数）
        if has_tip:
            if gt_x is None or gt_y is None:
                false_positive_count += 1
            elif distance is not None and distance > 4:
                false_positive_count += 1

        if (gt_x is not None and gt_y is not None) and not has_tip:
            miss_count += 1

        vis_path = out_img_dir / f"{img_name}_vis.jpg"
        draw_results(img, dets, gts, str(vis_path))

        per_image_records.append(
            {
                "文件名": img_name,
                "X坐标": det_x if det_x is not None else np.nan,
                "Y坐标": det_y if det_y is not None else np.nan,
                "是否有针尖": has_tip,
                "X_ture": gt_x if gt_x is not None else np.nan,
                "Y_true": gt_y if gt_y is not None else np.nan,
                "距离": distance if distance is not None else np.nan,
            }
        )

    df = pd.DataFrame(per_image_records)

    # 计算统计量
    valid_distances = [d for d in all_distances if d is not None and not np.isnan(d)]
    mean_err = np.mean(valid_distances) if valid_distances else np.nan
    max_err = np.max(valid_distances) if valid_distances else np.nan

    # 追加统计行（平均误差、最大误差、误检个数、漏检个数）
    stats_rows = pd.DataFrame(
        [
            {
                "文件名": "平均误差",
                "X坐标": "",
                "Y坐标": "",
                "是否有针尖": "",
                "X_ture": "",
                "Y_true": "",
                "距离": mean_err,
            },
            {
                "文件名": "最大误差",
                "X坐标": "",
                "Y坐标": "",
                "是否有针尖": "",
                "X_ture": "",
                "Y_true": "",
                "距离": max_err,
            },
            {
                "文件名": "误检个数",
                "X坐标": "",
                "Y坐标": "",
                "是否有针尖": "",
                "X_ture": "",
                "Y_true": "",
                "距离": false_positive_count,
            },
            {
                "文件名": "漏检个数",
                "X坐标": "",
                "Y坐标": "",
                "是否有针尖": "",
                "X_ture": "",
                "Y_true": "",
                "距离": miss_count,
            },
        ]
    )
    df = pd.concat([df, stats_rows], ignore_index=True)

    # 保存Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "TipEvaluation"

    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        if col_name in ["文件名", "X坐标", "Y坐标", "X_ture", "Y_true"]:
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = 12

    wb.save(out_excel)
    print(f"处理完成！结果保存在: {output_folder}")
    if valid_distances:
        print(f"平均 tip 误差: {mean_err:.2f} 像素")
        print(f"最大 tip 误差: {max_err:.2f} 像素")
    else:
        print("未找到任何有效匹配。")
    print(f"误检个数: {false_positive_count}")
    print(f"漏检个数: {miss_count}")


if __name__ == "__main__":
    main()
